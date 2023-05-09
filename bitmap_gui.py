
import sys
from pathlib import Path
from PIL import Image, ImageTk
from pprint import pprint

import tkinter as tk
from tkinter import END,Label, filedialog as fd
from tkinter import Button
import openpyxl
import numpy as np


from encoder import ImgToMonomer, rgb2hex
from colordepth import ReduceBitsPerColorChannel

from copy import copy

debug = True

class application:
    def __init__(
        self,
        window_width: int = 1600,
        window_height: int = 900) -> None:

        self.root = tk.Tk()
        self.root.title('Bitmap Encoder')
        self.raw_image = None
        self.raw_scaled = None
        self.im1 = None

        # Set x and y coordinates for the Tk root window
        ws = self.root.winfo_screenwidth() # width of the screen
        hs = self.root.winfo_screenheight() # height of the screen
        x = (ws/2) - (window_width/2)
        y = (hs/2) - (window_height/2)

        # Set the dimensions of the screen and where the window is placed
        self.root.geometry('%dx%d+%d+%d' % (window_width, window_height, x, y))

        # Buttons
        button_frame = tk.Frame(self.root)
        button_frame.grid(row = 0, column = 0)
        Button(button_frame, text='Exit Application', width=11, command=self.ExitApplication).grid(row=0, column=0)
        Button(button_frame, text='Select bitmap', fg='#000000', width=11, command=self.PickImageFile, bg='brown').grid(row=0, column=1)
        Button(button_frame, text='PrintRGB', fg='#000000', width=11,command=self.PrintRGB, bg='brown').grid(row=1, column=0)
        Button(button_frame, text='Export RGB', fg='#000000', width=11,command=self.ExportExcel, bg='brown').grid(row=1, column=1)

        # Sliders
        slider_frame = tk.Frame(self.root)
        slider_frame.grid(row=1, column = 0, sticky=tk.W)

        # Entry boxes
        self.bases = tk.Entry(slider_frame, width=4)
        self.bases.grid(row=3, column=0)
        self.bases.insert(0, str(4))
        Label(slider_frame, text='# Monomers').grid(row=4, column=0)

        bit_depth = tk.StringVar()
        bit_depth_entry = tk.Entry(slider_frame, width=4, textvariable=bit_depth)
        bit_depth_entry.grid(row=3, column=1)
        bit_depth_entry.insert(0, str(8))
        bit_depth_entry.bind('<Return>', (lambda _: self._ChangeColorDepth(bit_depth_entry)))
        Label(slider_frame, text='Color Depth').grid(row=4, column=1)

        # Text output
        self.output_text = tk.Text(master=self.root)
        self.output_text.grid(row=5, column=0, sticky=tk.NSEW)
        self.root.grid_rowconfigure(5, weight=1)
        sys.stdout = Redirect(self.output_text)

        # Images
        image_frame = tk.Frame(self.root, background='green', height=400, width=400)
        image_frame.grid(row=5, column=1)
        self.root.grid_columnconfigure(1, weight=5)

        self.left_im = Label(image_frame, image=self.im1)

        self.left_im.grid(row=0, column=0)

        self.root.mainloop()

    def ExportExcel(self):
        if self.im1 == None:
            MsgBox = tk.messagebox.showerror('Could not print RBG','You must have an active image\nto print the RGB values', icon = 'warning')
            return 0

        f = fd.asksaveasfilename(defaultextension='.xlsx')


        img = self.image_to_save
        # Check if a file was selected
        if f == '':
            return 1
        
        print(f)

        wb = openpyxl.Workbook()
        wb.create_sheet('encodings')
        encoding_sheet = wb['encodings']

        # Testing
        sheet = wb["Sheet"]

        # Get the encoding
        base = self.GetEncodingBase()
        encoding = ImgToMonomer(img, n_monomers=base, print_data=False)

        # Get the raw pixel data for coloring
        img = img.convert('RGB')
        pixels = np.array(img)

        for row_num, row in enumerate(pixels):
            for column_num, column in enumerate(row):
                # Get color in aRGB
                color = 'FF' + rgb2hex(column[0], column[1], column[2])[1:] # Remove the hashmark

                # Find the cell we're looking at
                cell = sheet.cell(row = row_num + 1, column = column_num + 1)
                encoding_cell = encoding_sheet.cell(row = row_num + 1, column = column_num + 1)

                # Change the value to encoding or the rgb
                cell.value = ' '.join(str(x) for x in column)
                encoding_cell.value = ''.join(str(x) for x in encoding[row_num][column_num]) # No space because space has meaning for encodings

                new_style = copy(cell.font)
                new_style.color = color
                cell.font = new_style
                encoding_cell.font = new_style

                cell.fill = openpyxl.styles.fills.GradientFill(type='linear', stop=[color, color])
                encoding_cell.fill = openpyxl.styles.fills.GradientFill(type='linear', stop=[color, color])
                cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
                encoding_cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT

                sheet.column_dimensions[openpyxl.utils.get_column_letter(column_num + 1)].width = 3
                encoding_sheet.column_dimensions[openpyxl.utils.get_column_letter(column_num + 1)].width = 3

            #sheet.row_dimensions[row_num + 1].height = 3


        wb.save(f)

    def _ChangeColorDepth(self, string_variable):
        '''
        Changes the color depth of the images.
        Callback function for getting a str from an Entry box
        '''
        if debug: print(string_variable.get())
        sv = int(''.join([char for char in string_variable.get() if str(char).isnumeric()]))

        # Set the original image to the reduced color depth
        self.im1 = ReduceBitsPerColorChannel(self.raw_scaled, sv)
        self.image_to_save = ReduceBitsPerColorChannel(self.raw_image, sv)
            
        self.UpdatePixelatedImage(event=None)
        self._DrawImages(event=None)

    def UpdatePixelatedImage(self, event):
        pixel_art_pil = self.GetPixelArtDisplayImage(self.im1, self.resample, (1600,2400))

        #self.UpdateImageEnhancements(None)
        self._DrawImages(None)

    def ExitApplication(self):
        MsgBox = tk.messagebox.askquestion ('Exit Application','Are you sure you want to exit the application', icon = 'warning')
        if MsgBox == 'yes':
            self.root.destroy()

    def PickImageFile(self):
        '''
        Select a raw image file for pixelation
        '''
        f = fd.askopenfilename()

        # Check if a file was selected
        if f == '':
            return 1

        im1 = Image.open(f)

        # Raw image for whatever purpose
        self.raw_image = Image.open(f)
        self.image_to_save = self.raw_image.copy()

        # Current size of the image
        size = (im1.width, im1.height)
        # Get scale factor. This sets the maximum sizes of the images
        #TODO Make the scaling of the image frame dynamic
        scale_width = size[0] / 500
        scale_height = size[1] / 900

        scale = max([scale_height, scale_width])
        # Rescale draw width and height
        size = (int(size[0] / scale), int(size[1] / scale))
        # Set the PIL im1 to the appropriate size
        self.im1 = im1.resize(size, resample=Image.NEAREST)
        # Raw_scaled
        self.raw_scaled = im1.resize(size, resample = Image.NEAREST)
    
        self._DrawImages(None)

    def _DrawImages(self, event):
        '''
        Draws the images. Should be called whenever changing
        self.im1 or self.im2
        '''
        self.tk_im1 = ImageTk.PhotoImage(self.im1)
        self.left_im.config(image=self.tk_im1)

    def GetEncodingBase(self) -> int:
        base = self.bases.get()

        if base == '':
            base = 4
        else:
            try:
                base = int(''.join([str(x) for x in base if x.isdigit()]))
            except:
                base = 4
        if base <= 1:
            base = 4
        
        if base == 4:
            self.bases.delete(0, END)
            self.bases.insert(0, str(4))

        return int(base)

    def PrintRGB(self):
        if self.im1 == None:
            MsgBox = tk.messagebox.showerror('Could not print RBG','You must have an active image\nto print the RGB values', icon = 'warning')
            return 0
        img = self.image_to_save
        
        self.output_text.delete('1.0', END)

        base = self.GetEncodingBase()

        encodings = ImgToMonomer(img, n_monomers=base)

        pprint(encodings)

    def SaveImage(self):
        if self.im1 == None:
            MsgBox = tk.messagebox.showerror('Could not print RBG','You must have an active image\nto print the RGB values', icon = 'warning')
            return 0

        # Get the resampled image for saving
        resample = (self.GetScaleX(None), self.GetScaleY(None))
        img = self.im1.resize(resample)

        f = Path(fd.asksaveasfilename())

        # Do some filename checks
        if f == '':
            return 1

        if f.suffix == '' or f.suffix != '.png':
            f = Path(str(f.absolute()) + '.png')

        img.save(str(f.resolve()))


class Redirect():
    '''
    Class for redirecting terminal output to the text widget
    '''
    def __init__(self, widget):
        self.widget = widget

    def write(self, text):
        self.widget.insert(END, text)
    
    def flush(self):
        pass

if __name__ == "__main__":
    application()